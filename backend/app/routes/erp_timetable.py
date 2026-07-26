# ============================================================
# SmartAttend — ERP Weekly Timetable API (v13)
#
# GET    /api/erp/timetable              — Admin slot grid / list
# POST   /api/erp/timetable/slot         — Create or update single slot
# DELETE /api/erp/timetable/slot/{id}    — Delete slot
# POST   /api/erp/timetable/duplicate    — Duplicate timetable (section A -> B)
# GET    /api/erp/timetable/student      — Student schedule (today + weekly)
# GET    /api/erp/timetable/teacher      — Teacher schedule (today + weekly)
# GET    /api/erp/timetable/current-period — Current period auto-fill
# GET    /api/erp/timetable/export       — Export Excel
# ============================================================

import io
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import get_current_admin, get_current_student, get_current_faculty
from app.models.models import (
    Admin, Student, Faculty, Classroom,
    ErpDepartment, ErpSubject, PeriodTiming, WeeklyTimetableSlot
)

router = APIRouter(prefix="/api/erp/timetable", tags=["ERP Timetable"])
logger = logging.getLogger(__name__)


# ─── Schemas ─────────────────────────────────────────────────

class SlotSaveRequest(BaseModel):
    department_id: int
    year: int
    section: str
    day_of_week: str
    period_timing_id: int
    erp_subject_id: Optional[int] = None
    faculty_id: Optional[int] = None
    classroom_id: Optional[int] = None
    class_type: str = "Theory"  # Theory | Lab | Elective | Tutorial | Break | Lunch | Free
    academic_year: Optional[str] = "2025-2026"
    semester: Optional[int] = 1


class DuplicateRequest(BaseModel):
    source_department_id: int
    source_year: int
    source_section: str
    target_department_id: int
    target_year: int
    target_section: str


# ─── Helpers ─────────────────────────────────────────────────

def _slot_dict(s: WeeklyTimetableSlot, now_time: str = "") -> dict:
    status = "upcoming"
    if s.period_timing and s.period_timing.start_time and s.period_timing.end_time:
        st = s.period_timing.start_time
        et = s.period_timing.end_time
        if st <= now_time < et:
            status = "active"
        elif et <= now_time:
            status = "completed"

    return {
        "id":               s.id,
        "department_id":    s.department_id,
        "department_name":  s.department.name if s.department else None,
        "department_short": s.department.short_name if s.department else None,
        "year":             s.year,
        "section":          s.section,
        "day_of_week":      s.day_of_week,
        "period_timing_id": s.period_timing_id,
        "period_label":     s.period_timing.label if s.period_timing else None,
        "start_time":       s.period_timing.start_time if s.period_timing else "",
        "end_time":         s.period_timing.end_time if s.period_timing else "",
        "order_index":       s.period_timing.order_index if s.period_timing else 0,
        "erp_subject_id":   s.erp_subject_id,
        "subject_name":     s.erp_subject.subject_name if s.erp_subject else None,
        "subject_code":     s.erp_subject.subject_code if s.erp_subject else None,
        "faculty_id":       s.faculty_id,
        "faculty_name":     s.faculty.name if s.faculty else None,
        "classroom_id":     s.classroom_id,
        "room_name":        s.classroom.room_name if s.classroom else None,
        "class_type":       s.class_type,
        "academic_year":    s.academic_year,
        "semester":         s.semester,
        "status":           status,
    }


# ─── Admin Endpoints ──────────────────────────────────────────

@router.get("", summary="Get timetable slot grid")
def get_timetable_grid(
    department_id: int = Query(...),
    year: int = Query(...),
    section: str = Query(...),
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == department_id,
        WeeklyTimetableSlot.year == year,
        WeeklyTimetableSlot.section == section,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    grid: Dict[str, List[dict]] = {d: [] for d in days}
    now = datetime.now().strftime("%H:%M")

    for s in slots:
        if s.day_of_week in grid:
            grid[s.day_of_week].append(_slot_dict(s, now))

    # Sort each day by order_index
    for d in grid:
        grid[d].sort(key=lambda x: x["order_index"])

    return {
        "department_id": department_id,
        "year": year,
        "section": section,
        "grid": grid,
        "total_slots": len(slots),
    }


@router.post("/slot", summary="Create or update timetable slot")
def save_slot(
    body: SlotSaveRequest,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    # Upsert slot by (department_id, year, section, day_of_week, period_timing_id)
    slot = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == body.department_id,
        WeeklyTimetableSlot.year == body.year,
        WeeklyTimetableSlot.section == body.section,
        WeeklyTimetableSlot.day_of_week == body.day_of_week,
        WeeklyTimetableSlot.period_timing_id == body.period_timing_id,
    ).first()

    if slot:
        slot.erp_subject_id = body.erp_subject_id
        slot.faculty_id = body.faculty_id
        slot.classroom_id = body.classroom_id
        slot.class_type = body.class_type
        slot.academic_year = body.academic_year
        slot.semester = body.semester
        slot.is_active = True
        slot.updated_at = datetime.now()
    else:
        slot = WeeklyTimetableSlot(
            department_id=body.department_id,
            year=body.year,
            section=body.section,
            day_of_week=body.day_of_week,
            period_timing_id=body.period_timing_id,
            erp_subject_id=body.erp_subject_id,
            faculty_id=body.faculty_id,
            classroom_id=body.classroom_id,
            class_type=body.class_type,
            academic_year=body.academic_year,
            semester=body.semester,
            is_active=True,
        )
        db.add(slot)

    db.commit()
    db.refresh(slot)
    return {"success": True, "slot": _slot_dict(slot)}


@router.delete("/slot/{slot_id}", summary="Delete timetable slot")
def delete_slot(
    slot_id: int,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    slot = db.query(WeeklyTimetableSlot).filter(WeeklyTimetableSlot.id == slot_id).first()
    if not slot:
        raise HTTPException(status_code=404, detail="Slot not found")
    db.delete(slot)
    db.commit()
    return {"success": True, "message": f"Slot {slot_id} deleted"}


@router.post("/duplicate", summary="Duplicate timetable from one section to another")
def duplicate_timetable(
    body: DuplicateRequest,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    source_slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == body.source_department_id,
        WeeklyTimetableSlot.year == body.source_year,
        WeeklyTimetableSlot.section == body.source_section,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    if not source_slots:
        raise HTTPException(status_code=404, detail="No source slots found to copy")

    # Clear target slots first
    db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == body.target_department_id,
        WeeklyTimetableSlot.year == body.target_year,
        WeeklyTimetableSlot.section == body.target_section,
    ).delete(synchronize_session=False)

    copied = 0
    for s in source_slots:
        new_slot = WeeklyTimetableSlot(
            department_id=body.target_department_id,
            year=body.target_year,
            section=body.target_section,
            day_of_week=s.day_of_week,
            period_timing_id=s.period_timing_id,
            erp_subject_id=s.erp_subject_id,
            faculty_id=s.faculty_id,
            classroom_id=s.classroom_id,
            class_type=s.class_type,
            academic_year=s.academic_year,
            semester=s.semester,
            is_active=True,
        )
        db.add(new_slot)
        copied += 1

    db.commit()
    return {"success": True, "copied": copied, "message": f"Successfully copied {copied} slots to target section"}


# ─── Student Schedule ─────────────────────────────────────────

@router.get("/student", summary="Get student timetable (auto-detected)")
def get_student_timetable(
    current_user: Student = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    today_day = datetime.now().strftime("%A")
    now_time = datetime.now().strftime("%H:%M")

    # Find department_id by matching name/short_name or default to 1st dept
    dept = db.query(ErpDepartment).filter(
        (ErpDepartment.short_name == current_user.department) |
        (ErpDepartment.name == current_user.department)
    ).first()
    dept_id = dept.id if dept else 1

    slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == dept_id,
        WeeklyTimetableSlot.year == current_user.year,
        WeeklyTimetableSlot.section == current_user.section,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    today_slots = [s for s in slots if s.day_of_week == today_day]
    today_schedule = [_slot_dict(s, now_time) for s in sorted(today_slots, key=lambda x: x.period_timing.order_index if x.period_timing else 0)]

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    weekly: Dict[str, List[dict]] = {d: [] for d in days}
    for s in slots:
        if s.day_of_week in weekly:
            weekly[s.day_of_week].append(_slot_dict(s, now_time))

    for d in weekly:
        weekly[d].sort(key=lambda x: x["order_index"])

    current = None
    upcoming = None
    for s in today_schedule:
        if s["status"] == "active":
            current = s
        elif s["status"] == "upcoming" and upcoming is None and s["class_type"] not in ("Break", "Lunch"):
            upcoming = s

    return {
        "today": today_schedule,
        "weekly": weekly,
        "current_period": current,
        "upcoming_period": upcoming,
        "today_day": today_day,
        "now": now_time,
        "student_info": {
            "department": current_user.department,
            "year": current_user.year,
            "section": current_user.section,
        }
    }


# ─── Teacher Schedule & Current Period ────────────────────────

@router.get("/teacher", summary="Get teacher timetable")
def get_teacher_timetable(
    current_user: Faculty = Depends(get_current_faculty),
    db: Session = Depends(get_db),
):
    today_day = datetime.now().strftime("%A")
    now_time = datetime.now().strftime("%H:%M")

    slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.faculty_id == current_user.id,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    today_slots = [s for s in slots if s.day_of_week == today_day]
    today_schedule = [_slot_dict(s, now_time) for s in sorted(today_slots, key=lambda x: x.period_timing.order_index if x.period_timing else 0)]

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    weekly: Dict[str, List[dict]] = {d: [] for d in days}
    for s in slots:
        if s.day_of_week in weekly:
            weekly[s.day_of_week].append(_slot_dict(s, now_time))

    for d in weekly:
        weekly[d].sort(key=lambda x: x["order_index"])

    current = None
    upcoming = None
    for s in today_schedule:
        if s["status"] == "active":
            current = s
        elif s["status"] == "upcoming" and upcoming is None and s["class_type"] not in ("Break", "Lunch"):
            upcoming = s

    auto_fill = None
    if current:
        auto_fill = {
            "department": current["department_short"],
            "year": current["year"],
            "section": current["section"],
            "subject_id": current["erp_subject_id"],
            "subject_name": current["subject_name"],
            "classroom_id": current["classroom_id"],
            "room": current["room_name"],
            "class_type": current["class_type"],
            "start_time": current["start_time"],
            "end_time": current["end_time"],
        }

    return {
        "today": today_schedule,
        "weekly": weekly,
        "current_period": current,
        "upcoming_period": upcoming,
        "auto_fill": auto_fill,
        "today_day": today_day,
        "now": now_time,
    }


@router.get("/current-period", summary="Get current timetable period for faculty auto-fill")
def get_current_period(
    current_user: Faculty = Depends(get_current_faculty),
    db: Session = Depends(get_db),
):
    today_day = datetime.now().strftime("%A")
    now_time = datetime.now().strftime("%H:%M")

    slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.faculty_id == current_user.id,
        WeeklyTimetableSlot.day_of_week == today_day,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    formatted = [_slot_dict(s, now_time) for s in slots]
    current = next((s for s in formatted if s["status"] == "active"), None)
    if not current:
        current = next((s for s in formatted if s["status"] == "upcoming" and s["class_type"] not in ("Break", "Lunch")), None)

    if not current:
        return {"found": False, "auto_fill": None}

    return {
        "found": True,
        "auto_fill": {
            "department": current["department_short"],
            "year": current["year"],
            "section": current["section"],
            "subject_id": current["erp_subject_id"],
            "subject_name": current["subject_name"],
            "subject_code": current["subject_code"],
            "classroom_id": current["classroom_id"],
            "room": current["room_name"],
            "class_type": current["class_type"],
            "period_number": current["order_index"],
            "start_time": current["start_time"],
            "end_time": current["end_time"],
        }
    }


# ─── Export Excel ─────────────────────────────────────────────

@router.get("/export", summary="Export weekly timetable as Excel")
def export_timetable(
    department_id: int = Query(...),
    year: int = Query(...),
    section: str = Query(...),
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    dept = db.query(ErpDepartment).filter(ErpDepartment.id == department_id).first()
    dept_name = dept.short_name if dept else f"Dept_{department_id}"

    slots = db.query(WeeklyTimetableSlot).filter(
        WeeklyTimetableSlot.department_id == department_id,
        WeeklyTimetableSlot.year == year,
        WeeklyTimetableSlot.section == section,
        WeeklyTimetableSlot.is_active == True,
    ).all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{dept_name}_Y{year}{section}"

    header = ["Day", "Period", "Start", "End", "Subject", "Code", "Faculty", "Room", "Type"]
    ws.append(header)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6C5CE7")

    for s in slots:
        ws.append([
            s.day_of_week,
            s.period_timing.label if s.period_timing else "",
            s.period_timing.start_time if s.period_timing else "",
            s.period_timing.end_time if s.period_timing else "",
            s.erp_subject.subject_name if s.erp_subject else "",
            s.erp_subject.subject_code if s.erp_subject else "",
            s.faculty.name if s.faculty else "",
            s.classroom.room_name if s.classroom else "",
            s.class_type,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Timetable_{dept_name}_Year{year}_{section}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
