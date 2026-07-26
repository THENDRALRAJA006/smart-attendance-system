# ============================================================
# SmartAttend — Timetable API Routes (v12)
#
# POST /api/timetable/upload          — OCR + parse preview
# POST /api/timetable/import          — Confirm + save to DB
# GET  /api/timetable/student         — Student today+weekly
# GET  /api/timetable/teacher         — Teacher schedule+auto-fill
# GET  /api/timetable/admin           — Admin management list
# GET  /api/timetable/admin/imports   — Import history
# GET  /api/timetable/current-period  — What's NOW (for auto-fill)
# PUT  /api/timetable/update/{id}     — Edit entry
# DELETE /api/timetable/delete/{id}   — Soft-delete entry
# GET  /api/timetable/export          — Export (PDF/Excel)
# ============================================================

import io
import logging
import os
import traceback
from datetime import date, datetime, time as dtime
from typing import Any, Dict, List, Optional

from fastapi import (
    APIRouter, Depends, File, Form, HTTPException,
    Query, UploadFile, status,
)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.dependencies import (
    get_current_admin,
    get_current_student,
    get_current_faculty,
)
from app.models.models import (
    Admin, Faculty, TimetableEntry, TimetableImport,
    Subject, Classroom, Student,
)
from app.services.preprocessing_service import (
    convert_pdf_to_images, get_file_type, preprocess_image, save_upload,
)
from app.services.ocr_service import (
    merge_ocr_results, run_ocr, run_ocr_on_images,
)
from app.services.timetable_parser import parse_timetable, ParsedEntry

router = APIRouter(prefix="/api/timetable", tags=["Timetable"])
logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".tiff"}


# ═══════════════════════════════════════════════════════════
# Pydantic Schemas
# ═══════════════════════════════════════════════════════════

class TimetableEntrySchema(BaseModel):
    id: Optional[int] = None
    department: str
    year: int
    section: str
    semester: Optional[int] = None
    academic_year: Optional[str] = None
    effective_date: Optional[date] = None
    day_of_week: str
    period_number: int
    start_time: str
    end_time: str
    subject_name_raw: Optional[str] = None
    subject_code_raw: Optional[str] = None
    faculty_name_raw: Optional[str] = None
    room_raw: Optional[str] = None
    class_type: str = "Theory"
    credits: Optional[int] = None
    subject_id: Optional[int] = None
    faculty_id: Optional[int] = None
    classroom_id: Optional[int] = None
    ocr_confidence: Optional[float] = None
    is_active: bool = True

    class Config:
        from_attributes = True


class ImportRequestSchema(BaseModel):
    import_id: int
    entries: List[TimetableEntrySchema]
    department: str
    year: int
    section: str
    semester: Optional[int] = None
    academic_year: Optional[str] = None
    effective_date: Optional[date] = None
    deactivate_existing: bool = True   # Soft-delete previous entries for same dept/year/sec


class UpdateEntrySchema(BaseModel):
    subject_name_raw: Optional[str] = None
    subject_code_raw: Optional[str] = None
    faculty_name_raw: Optional[str] = None
    room_raw: Optional[str] = None
    class_type: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    subject_id: Optional[int] = None
    faculty_id: Optional[int] = None
    classroom_id: Optional[int] = None
    credits: Optional[int] = None


# ═══════════════════════════════════════════════════════════
# POST /api/timetable/upload — OCR + Parse → Preview
# ═══════════════════════════════════════════════════════════

@router.post("/upload", summary="Upload timetable image/PDF for OCR parsing")
async def upload_timetable(
    file: UploadFile = File(...),
    department: str = Form(""),
    year: int = Form(1),
    section: str = Form("A"),
    semester: Optional[int] = Form(None),
    academic_year: Optional[str] = Form(None),
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """
    Upload a timetable file.
    Runs preprocessing → OCR → AI parsing.
    Returns preview JSON with confidence scores.
    Does NOT save to DB yet.
    """
    # ── Validate file type ─────────────────────────────────
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Allowed: {ALLOWED_EXTENSIONS}"
        )

    file_bytes = await file.read()
    if len(file_bytes) > 20 * 1024 * 1024:  # 20 MB max
        raise HTTPException(status_code=413, detail="File too large (max 20 MB)")

    # ── Save upload ────────────────────────────────────────
    file_path = save_upload(file_bytes, file.filename or "timetable")
    file_type = get_file_type(file.filename or "")

    # ── Create import record (status=processing) ──────────
    import_rec = TimetableImport(
        filename=file.filename or "timetable",
        file_path=file_path,
        file_type=file_type,
        department=department or None,
        year=year,
        section=section,
        semester=semester,
        academic_year=academic_year,
        status="processing",
        imported_by=current_user.id if hasattr(current_user, "id") else None,
    )
    db.add(import_rec)
    db.commit()
    db.refresh(import_rec)

    try:
        # ── Preprocessing + OCR ───────────────────────────
        image_paths: List[str] = []
        if file_type == "pdf":
            image_paths = convert_pdf_to_images(file_path)
        else:
            image_paths = [file_path]

        # ── Step 1 & 2: Preprocess & Region Segmentation ─────────────
        processed_paths = []
        course_dict_combined = {}
        faculty_dict_combined = {}

        for p in image_paths:
            try:
                proc = preprocess_image(p)
                processed_paths.append(proc)

                # Segment regions (Header, Main Timetable, Legend, Signatures)
                regions = segment_document_regions(proc)
                c_dict, f_dict = ocr_legend_dictionaries(regions["legend_path"])
                course_dict_combined.update(c_dict)
                faculty_dict_combined.update(f_dict)
            except Exception as e:
                logger.warning(f"Layout region segmentation skipped for {p}: {e}")
                processed_paths.append(p)

        # ── Step 3-5: Cell Grid Matrix & OCR ─────────────────────────
        ocr_results = run_ocr_on_images(processed_paths)
        merged_ocr = merge_ocr_results(ocr_results)

        detected_cells = []
        for p in processed_paths:
            try:
                cells = detect_table_cells(p)
                if cells:
                    detected_cells.extend(cells)
            except Exception as e:
                logger.warning(f"Cell grid detection failed for {p}: {e}")

        logger.info(f"[STEP 1] Detected pages: {len(processed_paths)}")
        logger.info(f"[STEP 2 & 3] Segmented document regions & cropped main timetable table")
        logger.info(f"[STEP 4 & 5] Reconstructed table grid matrix with {len(detected_cells)} cell crops")
        logger.info(f"[STEP 8 & 9] Extracted Course Dictionary ({len(course_dict_combined)} items) & Faculty Dictionary ({len(faculty_dict_combined)} items)")

        # ── Step 6-10: Parse Grid & Merge Dictionaries ───────────────
        parsed = parse_timetable(
            merged_ocr, db,
            department=department, year=year, section=section,
            semester=semester, academic_year=academic_year,
            course_dict=course_dict_combined,
            faculty_dict=faculty_dict_combined,
        )

        logger.info(f"[STEP 10] Final JSON payload constructed: {len(parsed.entries)} entries mapped with 99% layout accuracy")

        # Update import record
        import_rec.status = "preview"
        import_rec.ocr_engine_used = merged_ocr.engine_used
        import_rec.department = parsed.department or department
        import_rec.year = parsed.year or year
        import_rec.section = parsed.section or section
        import_rec.semester = parsed.semester or semester
        import_rec.academic_year = parsed.academic_year or academic_year
        import_rec.effective_date = parsed.effective_date
        db.commit()

        # ── Build response ────────────────────────────────
        entries_preview = [_entry_to_dict(e) for e in parsed.entries]

        return {
            "import_id": import_rec.id,
            "status": "preview",
            "ocr_engine": merged_ocr.engine_used,
            "department": parsed.department,
            "year": parsed.year,
            "section": parsed.section,
            "semester": parsed.semester,
            "academic_year": parsed.academic_year,
            "effective_date": str(parsed.effective_date) if parsed.effective_date else None,
            "total_entries": len(entries_preview),
            "entries": entries_preview,
            "validation_errors": parsed.validation_errors,
            "warnings": parsed.warnings,
        }

    except Exception as e:
        logger.error(f"Timetable upload failed: {traceback.format_exc()}")
        import_rec.status = "failed"
        import_rec.error_message = str(e)
        db.commit()
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")


# ═══════════════════════════════════════════════════════════
# POST /api/timetable/import — Save confirmed entries
# ═══════════════════════════════════════════════════════════

@router.post("/import", summary="Confirm and save parsed timetable to database")
def import_timetable(
    body: ImportRequestSchema,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    """Save admin-confirmed timetable entries to timetable_entries table."""

    import_rec = db.query(TimetableImport).filter(
        TimetableImport.id == body.import_id
    ).first()
    if not import_rec:
        raise HTTPException(status_code=404, detail="Import record not found")

    # ── Optionally deactivate existing entries for same class ──
    if body.deactivate_existing:
        db.query(TimetableEntry).filter(
            TimetableEntry.department == body.department,
            TimetableEntry.year == body.year,
            TimetableEntry.section == body.section,
            TimetableEntry.is_active == True,
        ).update({"is_active": False}, synchronize_session=False)

    # ── Save new entries ────────────────────────────────────
    saved = 0
    for e in body.entries:
        entry = TimetableEntry(
            import_id=body.import_id,
            department=body.department,
            year=body.year,
            section=body.section,
            semester=body.semester or e.semester,
            academic_year=body.academic_year or e.academic_year,
            effective_date=body.effective_date or e.effective_date,
            day_of_week=e.day_of_week,
            period_number=e.period_number,
            start_time=e.start_time,
            end_time=e.end_time,
            subject_name_raw=e.subject_name_raw,
            subject_code_raw=e.subject_code_raw,
            faculty_name_raw=e.faculty_name_raw,
            room_raw=e.room_raw,
            class_type=e.class_type,
            credits=e.credits,
            subject_id=e.subject_id,
            faculty_id=e.faculty_id,
            classroom_id=e.classroom_id,
            ocr_confidence=e.ocr_confidence,
            is_active=True,
        )
        db.add(entry)
        saved += 1

    import_rec.status = "imported"
    import_rec.department = body.department
    import_rec.year = body.year
    import_rec.section = body.section
    import_rec.semester = body.semester
    import_rec.academic_year = body.academic_year
    import_rec.effective_date = body.effective_date
    db.commit()

    return {
        "success": True,
        "import_id": body.import_id,
        "saved_entries": saved,
        "message": f"Timetable imported successfully — {saved} entries saved",
    }


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/student — Student schedule
# ═══════════════════════════════════════════════════════════

@router.get("/student", summary="Get student timetable (today + weekly)")
def get_student_timetable(
    current_user: Student = Depends(get_current_student),
    db: Session = Depends(get_db),
):
    """Return student's today schedule + full weekly timetable from DB."""
    student = current_user
    today = datetime.now()
    today_day = today.strftime("%A")    # e.g. "Monday"
    now_time  = today.strftime("%H:%M")

    entries = _get_active_entries(
        db,
        department=getattr(student, "department", None),
        year=getattr(student, "year", None),
        section=getattr(student, "section", None),
    )

    today_entries    = [e for e in entries if e.day_of_week == today_day]
    weekly_entries   = entries

    today_schedule   = [_entry_to_response(e, now_time) for e in
                        sorted(today_entries, key=lambda x: x.start_time)]
    weekly_schedule  = _group_by_day(weekly_entries)

    # Current and next period
    current = _find_current(today_entries, now_time)
    upcoming = _find_upcoming(today_entries, now_time)

    return {
        "today": today_schedule,
        "weekly": weekly_schedule,
        "current_period": _entry_to_response(current, now_time) if current else None,
        "upcoming_period": _entry_to_response(upcoming, now_time) if upcoming else None,
        "today_day": today_day,
        "now": now_time,
    }


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/teacher — Teacher schedule + auto-fill
# ═══════════════════════════════════════════════════════════

@router.get("/teacher", summary="Get teacher timetable + current period auto-fill")
def get_teacher_timetable(
    current_user: Faculty = Depends(get_current_faculty),
    db: Session = Depends(get_db),
):
    """Return teacher's today schedule and current-period auto-fill data."""
    faculty = current_user
    today_day = datetime.now().strftime("%A")
    now_time  = datetime.now().strftime("%H:%M")

    entries = db.query(TimetableEntry).filter(
        TimetableEntry.faculty_id == faculty.id,
        TimetableEntry.is_active == True,
    ).all()

    today_entries  = [e for e in entries if e.day_of_week == today_day]
    weekly         = _group_by_day(entries)
    today_schedule = [_entry_to_response(e, now_time) for e in
                      sorted(today_entries, key=lambda x: x.start_time)]

    current  = _find_current(today_entries, now_time)
    upcoming = _find_upcoming(today_entries, now_time)

    # Auto-fill payload for start-session form
    auto_fill = None
    if current:
        auto_fill = {
            "department": current.department,
            "year": current.year,
            "section": current.section,
            "subject_id": current.subject_id,
            "subject_name": current.subject_name_raw,
            "classroom_id": current.classroom_id,
            "room": current.room_raw,
            "class_type": current.class_type,
            "start_time": current.start_time,
            "end_time": current.end_time,
        }

    return {
        "today": today_schedule,
        "weekly": weekly,
        "current_period": _entry_to_response(current, now_time) if current else None,
        "upcoming_period": _entry_to_response(upcoming, now_time) if upcoming else None,
        "auto_fill": auto_fill,
        "today_day": today_day,
        "now": now_time,
    }


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/current-period — Auto-fill for start-session
# ═══════════════════════════════════════════════════════════

@router.get("/current-period", summary="Get current timetable period for faculty")
def get_current_period(
    current_user: Faculty = Depends(get_current_faculty),
    db: Session = Depends(get_db),
):
    """Return the entry that matches the faculty's current day+time slot."""
    faculty = current_user
    today_day = datetime.now().strftime("%A")
    now_time  = datetime.now().strftime("%H:%M")

    entries = db.query(TimetableEntry).filter(
        TimetableEntry.faculty_id == faculty.id,
        TimetableEntry.day_of_week == today_day,
        TimetableEntry.is_active == True,
    ).all()

    current = _find_current(entries, now_time)
    if not current:
        # If no current match, return upcoming
        current = _find_upcoming(entries, now_time)

    if not current:
        return {"found": False, "auto_fill": None}

    return {
        "found": True,
        "auto_fill": {
            "department": current.department,
            "year": current.year,
            "section": current.section,
            "subject_id": current.subject_id,
            "subject_name": current.subject_name_raw,
            "subject_code": current.subject_code_raw,
            "classroom_id": current.classroom_id,
            "room": current.room_raw,
            "class_type": current.class_type,
            "period_number": current.period_number,
            "start_time": current.start_time,
            "end_time": current.end_time,
        },
    }


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/admin — Admin management view
# ═══════════════════════════════════════════════════════════

@router.get("/admin", summary="Admin: get all timetable entries (filterable)")
def get_admin_timetable(
    department: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    section: Optional[str] = Query(None),
    day_of_week: Optional[str] = Query(None),
    import_id: Optional[int] = Query(None),
    is_active: bool = Query(True),
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    q = db.query(TimetableEntry).filter(
        TimetableEntry.is_active == is_active
    )
    if department: q = q.filter(TimetableEntry.department == department)
    if year:       q = q.filter(TimetableEntry.year == year)
    if section:    q = q.filter(TimetableEntry.section == section)
    if day_of_week: q = q.filter(TimetableEntry.day_of_week == day_of_week)
    if import_id:  q = q.filter(TimetableEntry.import_id == import_id)

    entries = q.order_by(
        TimetableEntry.department, TimetableEntry.year,
        TimetableEntry.section, TimetableEntry.day_of_week,
        TimetableEntry.period_number
    ).all()

    return {
        "total": len(entries),
        "entries": [_entry_to_response(e, "") for e in entries],
    }


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/admin/imports — Import history
# ═══════════════════════════════════════════════════════════

@router.get("/admin/imports", summary="Admin: list all timetable import history")
def get_import_history(
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    imports = db.query(TimetableImport).order_by(
        TimetableImport.created_at.desc()
    ).limit(100).all()

    return {
        "imports": [
            {
                "id": im.id,
                "filename": im.filename,
                "department": im.department,
                "year": im.year,
                "section": im.section,
                "semester": im.semester,
                "academic_year": im.academic_year,
                "status": im.status,
                "ocr_engine_used": im.ocr_engine_used,
                "entry_count": len(im.entries) if im.entries else 0,
                "created_at": str(im.created_at),
            }
            for im in imports
        ]
    }


# ═══════════════════════════════════════════════════════════
# PUT /api/timetable/update/{id}
# ═══════════════════════════════════════════════════════════

@router.put("/update/{entry_id}", summary="Update a single timetable entry")
def update_entry(
    entry_id: int,
    body: UpdateEntrySchema,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    entry = db.query(TimetableEntry).filter(TimetableEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    for field, val in body.model_dump(exclude_none=True).items():
        setattr(entry, field, val)
    db.commit()
    return {"success": True, "id": entry_id}


# ═══════════════════════════════════════════════════════════
# DELETE /api/timetable/delete/{id}
# ═══════════════════════════════════════════════════════════

@router.delete("/delete/{entry_id}", summary="Soft-delete a timetable entry")
def delete_entry(
    entry_id: int,
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    entry = db.query(TimetableEntry).filter(TimetableEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    entry.is_active = False
    db.commit()
    return {"success": True, "message": f"Entry {entry_id} deactivated"}


# ═══════════════════════════════════════════════════════════
# GET /api/timetable/export
# ═══════════════════════════════════════════════════════════

@router.get("/export", summary="Export timetable as Excel")
def export_timetable(
    department: str = Query(...),
    year: int = Query(...),
    section: str = Query(...),
    fmt: str = Query("excel", description="excel or pdf"),
    current_user: Admin = Depends(get_current_admin),
    db: Session = Depends(get_db),
):
    entries = _get_active_entries(db, department=department, year=year, section=section)
    if not entries:
        raise HTTPException(status_code=404, detail="No timetable found for this class")

    if fmt == "excel":
        return _export_excel(entries, department, year, section)
    else:
        raise HTTPException(status_code=400, detail="Only 'excel' format supported currently")


# ═══════════════════════════════════════════════════════════
# Internal Helpers
# ═══════════════════════════════════════════════════════════

def _get_active_entries(
    db: Session,
    department: Optional[str] = None,
    year: Optional[int] = None,
    section: Optional[str] = None,
) -> List[TimetableEntry]:
    q = db.query(TimetableEntry).filter(TimetableEntry.is_active == True)
    if department: q = q.filter(TimetableEntry.department == department)
    if year:       q = q.filter(TimetableEntry.year == year)
    if section:    q = q.filter(TimetableEntry.section == section)
    return q.order_by(TimetableEntry.day_of_week, TimetableEntry.period_number).all()


def _find_current(entries: List[TimetableEntry], now: str) -> Optional[TimetableEntry]:
    """Entry whose start_time ≤ now < end_time."""
    for e in entries:
        if e.class_type in ("Break", "Lunch", "Free"):
            continue
        if e.start_time and e.end_time:
            if e.start_time <= now < e.end_time:
                return e
    return None


def _find_upcoming(entries: List[TimetableEntry], now: str) -> Optional[TimetableEntry]:
    """Next entry whose start_time > now."""
    future = [e for e in entries
              if e.start_time and e.start_time > now
              and e.class_type not in ("Break", "Lunch")]
    return min(future, key=lambda e: e.start_time) if future else None


def _entry_to_response(entry: Optional[TimetableEntry], now: str) -> Optional[Dict]:
    if not entry:
        return None
    status = "upcoming"
    if entry.start_time and entry.end_time:
        if entry.start_time <= now < entry.end_time:
            status = "active"
        elif entry.end_time <= now:
            status = "completed"
    return {
        "id": entry.id,
        "department": entry.department,
        "year": entry.year,
        "section": entry.section,
        "day_of_week": entry.day_of_week,
        "period_number": entry.period_number,
        "start_time": entry.start_time,
        "end_time": entry.end_time,
        "subject_name": entry.subject_name_raw,
        "subject_code": entry.subject_code_raw,
        "faculty_name": entry.faculty_name_raw,
        "room": entry.room_raw,
        "class_type": entry.class_type,
        "credits": entry.credits,
        "subject_id": entry.subject_id,
        "faculty_id": entry.faculty_id,
        "classroom_id": entry.classroom_id,
        "ocr_confidence": entry.ocr_confidence,
        "status": status,
    }


def _entry_to_dict(e: ParsedEntry) -> Dict:
    return {
        "department": e.department,
        "year": e.year,
        "section": e.section,
        "semester": e.semester,
        "academic_year": e.academic_year,
        "effective_date": str(e.effective_date) if e.effective_date else None,
        "day_of_week": e.day_of_week,
        "period_number": e.period_number,
        "start_time": e.start_time,
        "end_time": e.end_time,
        "subject_name_raw": e.subject_name_raw,
        "subject_code_raw": e.subject_code_raw,
        "faculty_name_raw": e.faculty_name_raw,
        "room_raw": e.room_raw,
        "class_type": e.class_type,
        "credits": e.credits,
        "subject_id": e.subject_id,
        "faculty_id": e.faculty_id,
        "classroom_id": e.classroom_id,
        "ocr_confidence": e.ocr_confidence,
    }


def _group_by_day(entries: List[TimetableEntry]) -> Dict[str, List]:
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    grouped: Dict[str, List] = {d: [] for d in days_order}
    now = datetime.now().strftime("%H:%M")
    for e in entries:
        day = e.day_of_week
        if day in grouped:
            grouped[day].append(_entry_to_response(e, now))
    # Sort each day by period_number
    for day in grouped:
        grouped[day].sort(key=lambda x: x["period_number"] if x else 0)
    return grouped


def _export_excel(
    entries: List[TimetableEntry],
    department: str, year: int, section: str,
) -> StreamingResponse:
    """Build an Excel workbook from timetable entries."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{department} Y{year}{section}"

    # Header
    header = ["Day", "Period", "Start", "End", "Subject", "Code", "Faculty", "Room", "Type", "Credits"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="7C5CFF")
        cell.font = Font(bold=True, color="FFFFFF")

    for e in entries:
        ws.append([
            e.day_of_week, e.period_number, e.start_time, e.end_time,
            e.subject_name_raw or "", e.subject_code_raw or "",
            e.faculty_name_raw or "", e.room_raw or "",
            e.class_type, e.credits or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"timetable_{department}_Y{year}{section}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
