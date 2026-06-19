# ============================================================
# SmartAttend — Report Generation Service
# Excel, CSV, PDF export
# ============================================================

import io
import csv
import logging
from datetime import date
from typing import List
from sqlalchemy.orm import Session

from app.models.models import Attendance, Student, Subject, Classroom, Faculty, FacultySubject

logger = logging.getLogger(__name__)


def _get_attendance_data(
    db: Session,
    faculty_id: int,
    period: str,
    session_id: int | None = None,
    subject_id: int | None = None,
) -> List[dict]:
    """
    Fetch attendance records for the given faculty and filters.
    
    Returns:
        List of dicts with attendance info
    """
    from datetime import datetime, timedelta
    from sqlalchemy import and_

    today = date.today()
    if period == "daily":
        start_date = today
    elif period == "weekly":
        start_date = today - timedelta(days=7)
    else:  # monthly
        start_date = today.replace(day=1)

    query = (
        db.query(
            Attendance,
            Student.name.label("student_name"),
            Student.reg_no,
            Subject.subject_name,
            Subject.subject_code,
            Classroom.room_name,
            Faculty.name.label("faculty_name"),
        )
        .join(Student, Attendance.student_id == Student.id)
        .join(Subject, Attendance.subject_id == Subject.id)
        .join(Classroom, Attendance.classroom_id == Classroom.id)
        .join(Faculty, Subject.faculty_id == Faculty.id)
        .join(FacultySubject, FacultySubject.subject_id == Subject.id)
        .filter(
            FacultySubject.faculty_id == faculty_id,
            Attendance.date >= start_date,
        )
    )

    if session_id:
        query = query.filter(Attendance.session_id == session_id)
    if subject_id:
        query = query.filter(Attendance.subject_id == subject_id)

    # Deduplicate (junction table may cause duplicates for multi-faculty subjects)
    query = query.distinct()

    rows = query.order_by(Attendance.date.desc(), Attendance.time.desc()).all()

    return [
        {
            "id": r.Attendance.id,
            "student_name": r.student_name,
            "reg_no": r.reg_no,
            "subject": r.subject_name,
            "subject_code": r.subject_code,
            "classroom": r.room_name,
            "faculty_name": r.faculty_name,
            "date": r.Attendance.date.isoformat(),
            "time": r.Attendance.time,
            "status": r.Attendance.status,
            "rssi": r.Attendance.rssi,
            "face_confidence": r.Attendance.face_confidence,
        }
        for r in rows
    ]


def generate_csv(data: List[dict]) -> bytes:
    """Generate CSV bytes from attendance data."""
    output = io.StringIO()
    if not data:
        return b"No data available"

    fieldnames = [
        "id", "student_name", "reg_no", "subject", "subject_code",
        "classroom", "faculty_name", "date", "time", "status", "rssi", "face_confidence"
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue().encode("utf-8")


def generate_excel(data: List[dict]) -> bytes:
    """Generate Excel (.xlsx) bytes from attendance data."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # ─── Header style ────────────────────────────────────────
    header_fill = PatternFill(
        start_color="7C5CFF", end_color="7C5CFF", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = [
        "ID", "Student Name", "Reg No", "Subject", "Subject Code",
        "Classroom", "Faculty", "Date", "Time", "Status", "RSSI (dBm)", "Face Confidence (%)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ─── Data rows ───────────────────────────────────────────
    for row_idx, row in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value=row["student_name"])
        ws.cell(row=row_idx, column=3, value=row["reg_no"])
        ws.cell(row=row_idx, column=4, value=row["subject"])
        ws.cell(row=row_idx, column=5, value=row.get("subject_code", ""))
        ws.cell(row=row_idx, column=6, value=row["classroom"])
        ws.cell(row=row_idx, column=7, value=row.get("faculty_name", ""))
        ws.cell(row=row_idx, column=8, value=row["date"])
        ws.cell(row=row_idx, column=9, value=row["time"])
        ws.cell(row=row_idx, column=10, value=row["status"].upper())
        ws.cell(row=row_idx, column=11, value=row["rssi"])
        ws.cell(
            row=row_idx, column=12,
            value=f"{row['face_confidence']:.1f}" if row["face_confidence"] else "N/A"
        )

        # Color status cell
        status_cell = ws.cell(row=row_idx, column=10)
        if row["status"] == "present":
            status_cell.fill = PatternFill(
                start_color="00E676", end_color="00E676", fill_type="solid"
            )
        elif row["status"] == "absent":
            status_cell.fill = PatternFill(
                start_color="FF5252", end_color="FF5252", fill_type="solid"
            )

    # ─── Auto-size columns ───────────────────────────────────
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf(data: List[dict], title: str = "Attendance Report") -> bytes:
    """Generate PDF bytes from attendance data."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        raise RuntimeError("reportlab not installed")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        "title",
        parent=styles["Title"],
        textColor=colors.HexColor("#7C5CFF"),
        fontSize=18,
        spaceAfter=12,
    )
    elements.append(Paragraph(f"SmartAttend — {title}", title_style))
    elements.append(Paragraph(
        f"Generated: {date.today().strftime('%d %B %Y')}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 0.5 * cm))

    # Table
    headers = ["#", "Student", "Reg No", "Subject", "Code", "Classroom", "Faculty", "Date", "Time", "Status", "RSSI"]
    table_data = [headers]
    for row in data:
        table_data.append([
            str(row["id"]),
            row["student_name"],
            row["reg_no"],
            row["subject"],
            row.get("subject_code", ""),
            row["classroom"],
            row.get("faculty_name", ""),
            row["date"],
            row["time"],
            row["status"].upper(),
            str(row["rssi"] or ""),
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C5CFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0EEFF")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
    ]))
    elements.append(t)
    doc.build(elements)
    return buffer.getvalue()
